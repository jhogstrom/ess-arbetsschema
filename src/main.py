import argparse
import datetime
import os
from typing import Dict, Optional

import openpyxl
import pandas as pd
from dotenv import load_dotenv
from openpyxl.styles import Alignment, Border, Side
from pptx import Presentation
from pptx.dml.color import RGBColor

from helpers import FileHelper, color_boats, setup_logger

# Load .env file
load_dotenv()

# Get the defaults
default_file = os.getenv("REPORT_FILE")
default_date = os.getenv("REPORT_DATE")
default_template = os.getenv("TEMPLATE")
default_outdir = os.getenv("OUTDIR")


def parseargs():
    # Parse command line arguments
    parser = argparse.ArgumentParser()
    parser.add_argument("-f", "--file", default=default_file, help="Excel file to read")
    parser.add_argument(
        "-d", "--date", default=default_date, help="Date to generate report for"
    )
    parser.add_argument(
        "-t", "--template", default=default_template, help="Template file to fill in"
    )
    parser.add_argument(
        "-o", "--outdir", default=default_outdir, help="Filename to write the output to"
    )
    parser.add_argument(
        "--header", default="Schema ESS", help="Name of worksheet header"
    )
    parser.add_argument(
        "--mapfile", default="varvskarta*.pptx", help="Map file (powerpoint)"
    )
    return parser.parse_args()


def row_filter(row, report_date, schedule_name, data_settings: Dict[str, str]) -> bool:
    if schedule_name.upper() != row[data_settings["schedule_column"]].upper():
        return False
    if row[data_settings["date_column"]] != report_date:
        return False
    if pd.isna(row[data_settings["name_column"]]):
        return False
    return True


def get_dates(schedule: pd.DataFrame, schedule_name: str) -> list:
    year = datetime.datetime.now().year
    result = {
        row["Datum"]
        for _, row in schedule.iterrows()
        if datetime.datetime.strptime(row["Datum"], "%Y-%m-%d").year == year
        and schedule_name.upper() in row["Schema"].upper()
    }
    return sorted(result)


def make_report(
    *,
    date: str,
    schedule: pd.DataFrame,
    output_filename: str,
    map_output_filename: str,
    template: str,
    header: str,
    map_pptx: Optional[Presentation] = None,
    data_settings: dict,
) -> int:
    logger.info(f"Generating report for {date}")

    boatrows = sorted(
        [
            _
            for i, _ in schedule.iterrows()
            if row_filter(_, date, data_settings["boat_schedule"], data_settings)
        ],
        key=lambda x: x["Pass tid"],
    )

    work_rows = sorted(
        [
            _
            for i, _ in schedule.iterrows()
            if row_filter(_, date, data_settings["work_schedule"], data_settings)
        ],
        key=lambda x: x["Pass tid"],
    )

    # Load the template Excel file
    wb = openpyxl.load_workbook(template)

    # Select the sheet where you want to add the matchrows
    sheet = wb["Sheet1"]  # Replace 'Sheet1' with the name of your sheet

    # Specify the starting row and column
    start_row = 5

    # Define the border
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    def add_cell(sheet, row, col, value, wrap_text: bool = False):
        logger.debug(f"\tAdding cell {row}, {col}: {value}")
        cell = sheet.cell(row=row, column=col, value=value)
        cell.border = thin_border
        cell.alignment = Alignment(wrap_text=True)

    boats = []
    # Write the matchrows to the sheet
    sheet.insert_rows(start_row, len(boatrows))
    result = len(boatrows)
    for i, row in enumerate(boatrows, start=start_row):
        add_cell(sheet, i, 1, row["Pass tid"])
        # Medlem (fullt namn) pattern: "<namn> (<medlemsnummer>)"
        namn = row["Medlem (fullt namn)"].split("(")
        # Medlemsnummer
        id = namn[1][:-1]
        add_cell(sheet, i, 2, id)
        boats.append(id)
        # Medlemsnamn
        add_cell(sheet, i, 3, namn[0].strip())
        add_cell(sheet, i, 4, str(int(row["Mobil"])))
        add_cell(sheet, i, 5, row["Plats"])
        add_cell(sheet, i, 6, row["Modell"])
        kommentar = (
            row["Kommentar medlem"] if not pd.isna(row["Kommentar medlem"]) else ""
        )
        add_cell(sheet, i, 7, kommentar, wrap_text=True)
        esk = (
            "ESK: " + row["inställningESK"]
            if not pd.isna(row["inställningESK"])
            else None
        )
        dusk1 = (
            "DUSK1: " + row["inställningDUSK"]
            if not pd.isna(row["inställningDUSK"])
            else None
        )
        dusk2 = (
            "DUSK2: " + row["InställningDUSK2"]
            if not pd.isna(row["InställningDUSK2"])
            else None
        )
        settings = ", ".join(_ for _ in [esk, dusk1, dusk2] if _ is not None)
        add_cell(sheet, i, 8, settings)

    sheet.insert_rows(start_row + len(boatrows) + 4, len(work_rows))
    for i, row in enumerate(work_rows, start=start_row + len(boatrows) + 4):
        add_cell(sheet, i, 1, row["Pass tid"])
        namn = row["Medlem (fullt namn)"].split("(")
        add_cell(sheet, i, 2, namn[1][:-1])
        add_cell(sheet, i, 3, namn[0].strip())
        add_cell(sheet, i, 4, str(int(row["Mobil"])))

    sheet.cell(1, 1, f"{header} {date}")
    sheet.cell(
        row=1, column=7, value=datetime.datetime.now().strftime("%Y-%m-%d %H:%M")
    )

    # Save the workbook
    wb.save(output_filename)
    logger.info(f"Report written to '{output_filename}'")

    if map_pptx is not None:
        slide = map_pptx.slides[0]
        color_boats(
            slide, boats, RGBColor(255, 255, 26), "scheduled", logger, terse=False
        )
        map_pptx.save(map_output_filename)
    logger.info(f"Map written to '{output_filename}'")
    logger.info(
        f"== Summary '{header} {date}': {len(boatrows)} Arbetspass: {len(work_rows)}"
    )
    return result


def make_output_filename(outdir: str, date: str, suffix: str) -> str:
    return os.path.join(outdir, f"Förarschema ESS {date}.{suffix}")


def ensure_delete(filename: str):
    if os.path.exists(filename):
        os.remove(filename)


def find_balances(
    schedule: pd.DataFrame,
    data_settings: dict,
    boat_schedule_name: str,
    work_schedule_name: str,
):
    # print(schedule.columns)

    schedules = {}

    for row in schedule.iterrows():
        d = row[1][data_settings["date_column"]]
        if (
            datetime.datetime.strptime(d, "%Y-%m-%d").date()
            < datetime.datetime.today().date()
        ):
            continue
        key = d + " " + row[1][data_settings["schedule_time_column"]]
        if pd.isna(row[1][data_settings["name_column"]]):
            continue
        schema_type = row[1][data_settings["schedule_column"]]
        if schema_type not in [boat_schedule_name, work_schedule_name]:
            continue
        if key not in schedules:
            schedules[key] = {boat_schedule_name: [], work_schedule_name: []}

        schedules[key][schema_type].append(row[1][data_settings["email_column"]])

    for k, v in sorted(schedules.items(), key=lambda x: x[0]):
        if len(v[boat_schedule_name]) > 0 and len(v[work_schedule_name]) <= 1:
            print(
                k,
                f"saknas folk - {len(v[boat_schedule_name])} båtar {len(v[work_schedule_name])} medhjälpare",
            )
        if len(v[boat_schedule_name]) == 0 and len(v[work_schedule_name]) > 0:
            print(k, f"överbefolkat - 0 båtar {len(v[work_schedule_name])} medhjälpare")


def generate_reports(
    *,
    dates: list[str],
    schedule: pd.DataFrame,
    data_settings: dict,
    outdir: str,
    header: str,
    mapfile: str,
    template: str,
) -> Dict[str, int]:
    # Iterate over the dates. Generate a schedule for each date that is in the future
    # and delete the file if it is in the past
    stats = {}
    for d in dates:
        output_filename = make_output_filename(outdir, d, "xlsx")
        map_output_filename = make_output_filename(outdir, d, "pptx")

        if (
            datetime.datetime.strptime(d, "%Y-%m-%d").date()
            >= datetime.datetime.today().date()
        ):
            original_map_ppt = fh.read_pptx_file(
                fh.make_filename(mapfile, dirs=["templates"])
            )
            stats[d] = make_report(
                date=d,
                header=header,
                schedule=schedule,
                output_filename=output_filename,
                map_output_filename=map_output_filename,
                template=template,
                data_settings=data_settings,
                map_pptx=original_map_ppt,
            )
        else:
            ensure_delete(output_filename)
            ensure_delete(map_output_filename)
            logger.debug(f"**\n** Skipping passed date {d}\n**")
    return stats


if __name__ == "__main__":
    args = parseargs()
    logger = setup_logger("sched", "INFO")
    fh = FileHelper(logger)

    schedule_filename = fh.make_filename(args.file, dirs=["report", ".reports/reports"])
    logger.info(f"Reading schedule file '{schedule_filename}'")
    schedule = pd.read_excel(schedule_filename)
    BOAT_SCHEDULE = "Sjösättning 2025"
    WORK_SCHEDULE = "Arbetspass sjösättning 2025"

    data_settings = {
        "boat_schedule": BOAT_SCHEDULE,
        "work_schedule": WORK_SCHEDULE,
        "schedule_column": "Schema",
        "date_column": "Datum",
        "name_column": "Medlem (fullt namn)",
        "schedule_time_column": "Pass tid",
        "email_column": "Epost",
    }

    dates = get_dates(schedule, BOAT_SCHEDULE)
    if not os.path.exists(args.outdir):
        os.makedirs(args.outdir)
    stats = generate_reports(
        dates=dates,
        schedule=schedule,
        data_settings=data_settings,
        outdir=args.outdir,
        header=args.header,
        mapfile=args.mapfile,
        template=args.template,
    )

    find_balances(
        schedule,
        data_settings,
        boat_schedule_name=BOAT_SCHEDULE,
        work_schedule_name=WORK_SCHEDULE,
    )
    logger.info(f"Used schedule file '{schedule_filename}'")
    logger.info("Antal båtar per dag")
    for k, v in stats.items():
        logger.info(f"  {k}: {v}")
