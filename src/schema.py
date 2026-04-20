import argparse
import datetime
import json
import os
from typing import Any, Dict, List, Optional

import openpyxl
import pandas as pd
import requests
from dotenv import load_dotenv
from openpyxl.styles import Alignment, Border, Font, Side
from openpyxl.utils import get_column_letter
from pptx import presentation
from pptx.dml.color import RGBColor

from googleapi import get_google_sheet, get_sheet_titles
from helpers import FileHelper, color_boats, setup_logger

# Load .env file
load_dotenv()

# Get the defaults
default_file = os.getenv("REPORT_FILE")
default_date = os.getenv("REPORT_DATE")
default_outdir = os.getenv("OUTDIR")


def parseargs():
    # Parse command line arguments
    parser = argparse.ArgumentParser()
    parser.add_argument(
        "-f",
        "--file",
        default=default_file,
        metavar="<file>",
        help="Excel file to read.",
    )
    parser.add_argument(
        "-d", "--date", default=default_date, help="Date to generate report for."
    )
    parser.add_argument(
        "-o",
        "--outdir",
        default=default_outdir,
        metavar="<dir>",
        help="Directory to write the output to.",
    )
    parser.add_argument(
        "--header",
        default="Schema ESS",
        metavar="<str>",
        help="Name of worksheet header.",
    )
    parser.add_argument(
        "--mapfile",
        default="varvskarta*.pptx",
        metavar="<file>",
        help="Map file (powerpoint).",
    )
    parser.add_argument(
        "--driversheetid",
        metavar="<id>",
        default=os.getenv("DRIVERSCHEDULE"),
        help="Google Sheet ID to read the driver schedule from",
    )
    parser.add_argument(
        "--openweather_apikey",
        metavar="<key>",
        default=os.getenv("OPENWEATHER_API_KEY"),
        help="OpenWeather API key",
    )
    parser.add_argument(
        "--location",
        metavar="<location>",
        default=os.getenv("LOCATION"),
        help="Location for weather data",
    )
    return parser.parse_args()


def row_filter(row, report_date, schedule_name, data_settings: Dict[str, str]) -> bool:
    if schedule_name.upper() != row[data_settings["schedule_column"]].upper():
        return False
    if row[data_settings["date_column"]] != report_date:
        return False
    if pd.isna(row[data_settings["name_column"]]):
        return False
    return True


def get_dates(schedule: pd.DataFrame, schedule_name: str) -> list:
    year = datetime.datetime.now().year
    result = {
        row["Datum"]
        for _, row in schedule.iterrows()
        if datetime.datetime.strptime(row["Datum"], "%Y-%m-%d").year == year
        and schedule_name.upper() in row["Schema"].upper()
    }
    return sorted(result)


emails: Dict[str, List[str]] = {}
missing_foreman: List[str] = []
generated_files: Dict[str, List[str]] = {}


def remove_shapes(slide, shapes_to_remove, logger):
    for shape in slide.shapes:
        if shape.name in shapes_to_remove:
            logger.debug(f"Removing shape: {shape.name}")
            sp = shape._element
            sp.getparent().remove(sp)


def _get_rows(schedule: pd.DataFrame, date: str, schedule_name: str) -> list:
    return sorted(
        [
            _
            for i, _ in schedule.iterrows()
            if row_filter(_, date, schedule_name, data_settings)
        ],
        key=lambda x: x["Pass tid"],
    )


def _get_lat_long(location: str, api_key: str) -> tuple:
    CACHE_FILE = "location_cache.json"
    if os.path.exists(CACHE_FILE) and os.path.isfile(CACHE_FILE):
        with open(CACHE_FILE, "r", encoding="utf-8") as f:
            location_cache = json.load(f)
    else:
        location_cache = {}

    if location in location_cache:
        return tuple(location_cache[location])

    url = "https://api.openweathermap.org/geo/1.0/direct"
    params = {
        "q": location,
        "limit": 1,
        "appid": api_key,
    }
    response = requests.get(url, params=params)
    response.raise_for_status()
    data = response.json()
    if len(data) == 0:
        raise ValueError(f"Location '{location}' not found")
    location_cache[location] = (data[0]["lat"], data[0]["lon"])
    with open(CACHE_FILE, "w", encoding="utf-8") as f:
        json.dump(location_cache, f, indent=2, ensure_ascii=False)
    return location_cache[location]


def get_weather(date: str, location: str, api_key: str) -> Optional[dict]:
    if api_key is None:
        return None

    lat, lon = _get_lat_long(location, api_key)
    exclude = "current,minutely,hourly,alerts"

    endpoint = "onecall"
    # endpoint = "onecall/timemachine"
    url = f"https://api.openweathermap.org/data/3.0/{endpoint}"
    params = {
        "lat": lat,
        "lon": lon,
        "appid": api_key,
        "exclude": exclude,
        "units": "metric",
        # "lang": "sv, se",
    }
    # icon: https://openweathermap.org/img/wn/10d@2x.png
    if endpoint == "onecall/timemachine":
        dt = int(
            datetime.datetime.strptime(date, "%Y-%m-%d").replace(hour=12).timestamp()
        )
        params["dt"] = dt
    response = requests.get(url, params=params)
    response.raise_for_status()

    data = response.json()
    for _ in data.get("daily", []):
        if datetime.datetime.fromtimestamp(_["dt"]).strftime("%Y-%m-%d") == date:
            return _

    return None


def parse_weather(data: Optional[dict]) -> Optional[dict]:
    if data is None:
        return None
    temp = data.get("temp", {}).get("day")
    return {
        "temp": f"{int(round(temp))}C på dagen" if temp is not None else "-",
        "description": f"{data.get('summary')} ({data.get('weather', [{}])[0].get('description')}).",
        "icon": data.get("weather", [{}])[0].get("icon"),
        "wind_speed": f"{int(round(data.get('wind_speed', 0)))} ({int(round(data.get('wind_gust', 0)))}) m/s",
    }


def _make_excel_report(
    boatrows,
    work_rows,
    foreman_rows,
    filename: str,
    *,
    header: str,
    date: str,
    drivers: List[str],
    openweather_apikey: Optional[str] = None,
    location: Optional[str] = None,
) -> None:
    # Create a new workbook and select the active sheet
    wb = openpyxl.Workbook()
    sheet = wb.active
    if sheet is None:
        raise ValueError("No active sheet found in the workbook")
    sheet.title = "Schema"

    # Define the border
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    def add_cell(
        sheet,
        row,
        col,
        value,
        wrap_text: bool = False,
        *,
        width: Optional[float] = None,
        header: bool = False,
        border: bool = True,
        multiline: bool = False,
    ) -> Any:
        logger.debug(f"\tAdding cell {row}, {col}: {value}")
        cell = sheet.cell(row=row, column=col, value=value)
        if border:
            cell.border = thin_border
        cell.alignment = Alignment(wrap_text=wrap_text)
        if width is not None:
            sheet.column_dimensions[get_column_letter(col)].width = width
        if header:
            cell.font = Font(bold=True, size=13)
        if multiline:
            cell.alignment = Alignment(wrap_text=True, vertical="top")
        else:
            cell.alignment = Alignment(vertical="top")
        return cell

    # Add the header
    add_cell(sheet, 1, 1, f"{header} {date}", header=True, border=False)
    add_cell(
        sheet,
        1,
        7,
        datetime.datetime.now().strftime("%Y-%m-%d %H:%M"),
        header=True,
        border=False,
    )

    # Add the column headers
    add_cell(sheet, 3, 1, "Pass", width=12, header=True)
    add_cell(sheet, 3, 2, "#", width=6, header=True)
    add_cell(sheet, 3, 3, "Namn", width=17, header=True)
    add_cell(sheet, 3, 4, "Mobil", width=14, header=True)
    add_cell(sheet, 3, 5, "Plats", width=6, header=True)
    add_cell(sheet, 3, 6, "Båtmodell", width=20, header=True)
    add_cell(sheet, 3, 7, "Kommentar", width=30, header=True)
    add_cell(sheet, 3, 8, "Inställningar", width=20, header=True)
    # Specify the starting row and column
    next_row = 4
    for i, row in enumerate(boatrows, start=next_row):
        next_row += 1
        add_cell(sheet, i, 1, row["Pass tid"])
        # Medlem (fullt namn) pattern: "<namn> (<medlemsnummer>)"
        namn = row["Medlem (fullt namn)"].split("(")
        # Medlemsnummer
        id = int(namn[1][:-1])
        add_cell(sheet, i, 2, id)
        # Medlemsnamn
        add_cell(sheet, i, 3, namn[0].strip())
        add_cell(sheet, i, 4, " " + str(int(row["Mobil"])))
        plats = [_.strip() for _ in str(row["Plats"]).split(",")]
        add_cell(sheet, i, 5, ", ".join(set(plats)))
        add_cell(sheet, i, 6, row["Modell"])
        kommentar = (
            row["Kommentar medlem"] if not pd.isna(row["Kommentar medlem"]) else ""
        )
        add_cell(sheet, i, 7, kommentar, wrap_text=True)
        esk = (
            "ESK: " + row["inställningESK"]
            if not pd.isna(row["inställningESK"])
            else None
        )
        dusk1 = (
            "DUSK1: " + row["inställningDUSK"]
            if not pd.isna(row["inställningDUSK"])
            else None
        )
        dusk2 = (
            "DUSK2: " + row["InställningDUSK2"]
            if not pd.isna(row["InställningDUSK2"])
            else None
        )
        settings = "\n".join(_ for _ in [esk, dusk1, dusk2] if _ is not None)
        add_cell(sheet, i, 8, settings, multiline=True)

    next_row += 2
    add_cell(sheet, next_row, 1, "Arbetspass", header=True)
    add_cell(sheet, next_row, 2, "#", header=True)
    add_cell(sheet, next_row, 3, "Namn", header=True)
    add_cell(sheet, next_row, 4, "Mobil", header=True)

    add_cell(sheet, next_row, 6, "Förmanspass", header=True)
    add_cell(sheet, next_row, 7, "Namn", header=True)
    add_cell(sheet, next_row, 8, "Mobil", header=True)

    next_row += 1
    for i, row in enumerate(work_rows, start=next_row):
        add_cell(sheet, i, 1, row["Pass tid"])
        namn = row["Medlem (fullt namn)"].split("(")
        add_cell(sheet, i, 2, int(namn[1][:-1]))
        add_cell(sheet, i, 3, namn[0].strip())
        add_cell(sheet, i, 4, " " + str(int(row["Mobil"])))

    for i, row in enumerate(foreman_rows, start=next_row):
        add_cell(sheet, i, 6, row["Pass tid"])
        namn = row["Medlem (fullt namn)"].split("(")
        add_cell(sheet, i, 7, namn[0].strip())
        add_cell(sheet, i, 8, str(int(row["Mobil"])))

    if len(foreman_rows) == 0:
        logger.warning(f"No foreman found for {date}!")
        add_cell(sheet, next_row, 7, "INGEN FÖRMAN", border=False)

    next_row += len(foreman_rows) + 2
    add_cell(sheet, next_row, 6, "Förare", header=True)
    for i, row in enumerate(drivers, start=next_row + 1):
        add_cell(sheet, i, 6, row[1])  # Name

    if (
        openweather_apikey is not None
        and location is not None
        and (
            datetime.datetime.strptime(date, "%Y-%m-%d").date()
            - datetime.datetime.today().date()
        ).days
        < 3
    ):
        weather = parse_weather(get_weather(date, location, api_key=openweather_apikey))
        if not weather:
            logger.warning(f"No weather data found for {date} {location}")
        else:
            next_row += len(drivers) + 2
            add_cell(sheet, next_row, 6, "Väderprognos", header=True)
            add_cell(sheet, next_row, 7, location, header=True)
            next_row += 1
            add_cell(sheet, next_row, 6, "Väder", header=True)
            add_cell(sheet, next_row, 7, weather["description"])
            next_row += 1
            add_cell(sheet, next_row, 6, "Temp", header=True)
            add_cell(sheet, next_row, 7, weather["temp"])
            next_row += 1
            add_cell(sheet, next_row, 6, "Vind", header=True)
            add_cell(sheet, next_row, 7, weather["wind_speed"])

    # Save the workbook
    wb.save(filename)
    logger.info(f"Report written to '{filename}'")


def _save_emails(
    boatrows,
    work_rows,
    foreman_rows,
    drivers,
    filename: str,
    *,
    date: str,
    email_column: str,
) -> None:
    todays_emails = [_[email_column] for _ in boatrows if not pd.isna(_[email_column])]
    todays_emails.extend(
        [_[email_column] for _ in work_rows if not pd.isna(_[email_column])]
    )
    todays_emails.extend(
        [_[email_column] for _ in foreman_rows if not pd.isna(_[email_column])]
    )
    email_idx = drivers[0].index(email_column) if email_column in drivers[0] else -1
    if email_idx == -1:
        raise ValueError(f"Email column '{email_column}' not found in drivers sheet")
    # Add also all drivers for that date
    todays_emails.extend(_[email_idx] for _ in drivers[1:] if _[0] == date)
    # Write the email list to a file
    with open(filename, "w", encoding="utf-8") as f:
        for e in sorted(set(todays_emails)):
            f.write(e + "\n")
    logger.info(f"Email list written to '{filename}'")


def _save_powerpoint(
    boatrows, filename: str, map_pptx: Optional[presentation.Presentation]
) -> None:
    if map_pptx is None:
        logger.warning("No map PPTX provided, skipping map generation.")
        return

    boats = [int(_["Medlemsnr"]) for _ in boatrows]
    slide = map_pptx.slides[0]
    color_boats(slide, boats, RGBColor(255, 255, 26), "scheduled", logger, terse=False)
    shapes_to_remove = [
        "Anteckning 1",
        "Anteckning 2",
        "Anteckning 3",
    ]
    remove_shapes(slide, shapes_to_remove, logger)

    map_pptx.save(filename)
    logger.info(f"Map written to '{filename}'")


def make_report(
    *,
    date: str,
    schedule: pd.DataFrame,
    output_filename: str,
    map_output_filename: str,
    email_output_filename: str,
    drivers,
    header: str,
    map_pptx: Optional[presentation.Presentation] = None,
    data_settings: dict,
    openweather_apikey: Optional[str] = None,
    location: Optional[str] = None,
) -> int:
    logger.info(f"Generating report for {date}")

    boatrows = _get_rows(schedule, date, data_settings["boat_schedule"])
    work_rows = _get_rows(schedule, date, data_settings["work_schedule"])
    foreman_rows = _get_rows(schedule, date, data_settings["foreman_schedule"])

    _make_excel_report(
        boatrows,
        work_rows,
        foreman_rows,
        output_filename,
        header=header,
        date=date,
        drivers=[_ for _ in drivers if _[0] == date],
        openweather_apikey=openweather_apikey,
        location=location,
    )
    _save_emails(
        boatrows,
        work_rows,
        foreman_rows,
        drivers,
        email_output_filename,
        date=date,
        email_column=data_settings["email_column"],
    )
    _save_powerpoint(boatrows, map_output_filename, map_pptx)

    logger.info(
        f"== Summary '{header} {date}': {len(boatrows)} Arbetspass: {len(work_rows)}"
    )
    if len(foreman_rows) == 0:
        missing_foreman.append(date)

    # Records the files that have been generated
    generated_files[date] = [
        output_filename,
        map_output_filename,
        email_output_filename,
    ]

    return len(boatrows)


def make_output_filename(outdir: str, date: str, suffix: str) -> str:
    return os.path.join(outdir, f"Förarschema ESS {date}.{suffix}")


def ensure_delete(filename: str):
    if os.path.exists(filename):
        os.remove(filename)


def find_balances(
    schedule: pd.DataFrame,
    data_settings: dict,
    boat_schedule_name: str,
    work_schedule_name: str,
):
    # print(schedule.columns)

    schedules = {}

    for row in schedule.iterrows():
        d = str(row[1][data_settings["date_column"]])
        if (
            datetime.datetime.strptime(d, "%Y-%m-%d").date()
            < datetime.datetime.today().date()
        ):
            continue
        key = d + " " + row[1][data_settings["schedule_time_column"]]
        if pd.isna(str(row[1][data_settings["name_column"]])):
            continue
        schema_type = row[1][data_settings["schedule_column"]]
        if schema_type not in [boat_schedule_name, work_schedule_name]:
            continue
        if key not in schedules:
            schedules[key] = {boat_schedule_name: [], work_schedule_name: []}

        schedules[key][schema_type].append(row[1][data_settings["email_column"]])

    for k, v in sorted(schedules.items(), key=lambda x: x[0]):
        if len(v[boat_schedule_name]) > 0 and len(v[work_schedule_name]) <= 1:
            print(
                k,
                f"saknas folk - {len(v[boat_schedule_name])} båtar {len(v[work_schedule_name])} medhjälpare",
            )
        if len(v[boat_schedule_name]) == 0 and len(v[work_schedule_name]) > 0:
            print(k, f"överbefolkat - 0 båtar {len(v[work_schedule_name])} medhjälpare")


def generate_reports(
    *,
    dates: list[str],
    schedule: pd.DataFrame,
    data_settings: dict,
    drivers,
    outdir: str,
    header: str,
    mapfile: str,
    openweather_apikey: Optional[str] = None,
    location: Optional[str] = None,
) -> Dict[str, int]:
    # Iterate over the dates. Generate a schedule for each date that is in the future
    # and delete the file if it is in the past
    stats = {}
    for d in dates:
        output_filename = make_output_filename(outdir, d, "xlsx")
        map_output_filename = make_output_filename(outdir, d, "pptx")
        email_output_filename = make_output_filename(outdir, d, "email.txt")

        if (
            datetime.datetime.strptime(d, "%Y-%m-%d").date()
            >= datetime.datetime.today().date()
        ):
            original_map_ppt = fh.read_pptx_file(
                fh.make_filename(mapfile, dirs=["templates", ".reports/templates"])
            )
            stats[d] = make_report(
                date=d,
                header=header,
                schedule=schedule,
                drivers=drivers,
                output_filename=output_filename,
                map_output_filename=map_output_filename,
                email_output_filename=email_output_filename,
                data_settings=data_settings,
                map_pptx=original_map_ppt,
                openweather_apikey=openweather_apikey,
                location=location,
            )
        else:
            ensure_delete(output_filename)
            ensure_delete(map_output_filename)
            ensure_delete(email_output_filename)
            logger.debug(f"**\n** Skipping passed date {d}\n**")
    return stats


def get_drivers(sheet_id):
    if sheet_id is None:
        logger.warning("No driver schedule sheet ID provided.")
        return []
    logger.info(f"Reading driver schedule from Google Sheet ID '{sheet_id[:5]}...'.")
    return get_google_sheet(sheet_id, get_sheet_titles(sheet_id)[0])


if __name__ == "__main__":
    args = parseargs()
    logger = setup_logger("sched", os.getenv("DEBUG_LEVEL", "DEBUG"))
    fh = FileHelper(logger)

    schedule_filename = fh.make_filename(args.file, dirs=["report", ".reports/reports"])
    logger.info(f"Reading schedule file '{schedule_filename}'")
    schedule = pd.read_excel(schedule_filename)

    # # print 10 first rows and columns of the schedule for debugging
    # logger.debug(f"Schedule columns: {schedule.columns.tolist()}")
    # logger.debug(f"First 10 rows of the schedule:\n{schedule.head(10)}")

    activity = "sjösättning" if datetime.datetime.now().month <= 7 else "torrsättning"
    current_year = datetime.datetime.now().year
    BOAT_SCHEDULE = f"{activity.capitalize()} {current_year}"
    WORK_SCHEDULE = f"Arbetspass {activity} {current_year}"
    FOREMAN_SCHEDULE = f"Förmanspass till {activity} {current_year} (för styrelsen)"

    data_settings = {
        "boat_schedule": BOAT_SCHEDULE,
        "work_schedule": WORK_SCHEDULE,
        "foreman_schedule": FOREMAN_SCHEDULE,
        "schedule_column": "Schema",
        "date_column": "Datum",
        "name_column": "Medlem (fullt namn)",
        "schedule_time_column": "Pass tid",
        "email_column": "Epost",
    }

    dates = get_dates(schedule, BOAT_SCHEDULE)
    if not os.path.exists(args.outdir):
        os.makedirs(args.outdir)
    drivers = get_drivers(args.driversheetid)
    stats = generate_reports(
        dates=dates,
        schedule=schedule,
        data_settings=data_settings,
        drivers=drivers,
        outdir=args.outdir,
        header=args.header,
        mapfile=args.mapfile,
        openweather_apikey=args.openweather_apikey,
        location=args.location,
    )

    find_balances(
        schedule,
        data_settings,
        boat_schedule_name=BOAT_SCHEDULE,
        work_schedule_name=WORK_SCHEDULE,
    )
    logger.info(f"Used schedule file '{schedule_filename}'")
    logger.info("Antal båtar per dag")
    for k, v in stats.items():
        logger.info(f"  {k}: {v}")

    filedata = {
        "parent_folder_id": os.getenv("PARENT_FOLDER_ID", ""),
        "files": generated_files,
    }
    with open("stage/generated_files.json", "w", encoding="utf-8") as f:
        json.dump(filedata, f, indent=2, ensure_ascii=False)

    for d in missing_foreman:
        logger.warning(f"No foreman assigned for {d}!")
